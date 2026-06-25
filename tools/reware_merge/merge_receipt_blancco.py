# -*- coding: utf-8 -*-
"""
Re-Ware: merge product list (model + serial) + Blancco export.
Detects files by column layout (not filename). CSV and Excel both supported.
Double-click run_merge.bat in the folder that contains the input files.
"""
from __future__ import annotations

import re
import sys
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill

# --- Config ---
SCRIPT_DIR = Path(__file__).resolve().parent
SUPPORTED_EXTENSIONS = (".csv", ".xlsx")
PRODUCT_LIST_MAX_COLS = 6
BLANCCO_MIN_COLS = 8
MTM_LUT_FILE = SCRIPT_DIR / "mtm_lookup.csv"
OUTPUT_PREFIX = "MERGED import-ready"
OUTPUT_SKIP_PREFIXES = (OUTPUT_PREFIX.lower(), "merged import-ready")

SERIAL_S_PREFIX = re.compile(r"^S((?:PC|PF|GM|R)\w+)$", re.I)
SCAN_SERIAL_TAIL_RE = re.compile(r"(?P<serial>(?:PC|PF|GM|R)[A-Z0-9]{6,})$", re.I)
LENOVO_MTM_RE = re.compile(r"^\d{2}[A-Z0-9]{8}$", re.I)
PORTAL_SCAN_NAME_MARKERS = ("scanned stock", "portal")
GEN_RE = re.compile(r"gen\s*(\d+\w*)", re.I)
SYSVER_GEN_RE = re.compile(r"Gen(?:eration)?\s*(\d+\w*)", re.I)

BATTERY_TIER_THRESHOLD = 70

OUTPUT_COLUMNS = [
    "Serial",
    "MTM",
    "Model name",
    "System version",
    "Series",
    "Touch",
    "WAN",
    "Generation",
    "CPU",
    "RAM (GB)",
    "SSD type",
    "SSD size (GB)",
    "Battery (%)",
    "Battery display",
    "Battery tier",
    "Shop SKU",
    "GPU",
    "Mobo status",
    "Blancco date",
    "Manufacturer",
    "Price",
    "Status",
    "Failure reason",
]

FAIL_FILL = PatternFill(start_color="FFFFC7CE", end_color="FFFFC7CE", fill_type="solid")
HEADER_FILL = PatternFill(start_color="FFD9E1F2", end_color="FFD9E1F2", fill_type="solid")

def normalize_serial(value) -> str:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return ""
    s = str(value).strip().upper()
    if not s:
        return ""
    m = SERIAL_S_PREFIX.match(s)
    if m:
        return m.group(1).upper()
    return s

def normalize_mtm(value) -> str:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return ""
    return str(value).strip().upper()


def _is_portal_scan_noise(text: str) -> bool:
    s = (text or "").strip().upper()
    if not s or s == "NAN":
        return True
    if s in ("MONITORS", "QTY"):
        return True
    return "SAMSUNG" in s


def parse_portal_scan_row(col0, col1="") -> tuple[str, str]:
    """Split portal scanner rows (1S+MTM+SN glued, or MTM|SN columns)."""
    c0 = str(col0 or "").strip()
    c1 = str(col1 or "").strip()
    if _is_portal_scan_noise(c0):
        return "", ""
    if c0.upper().startswith("1S"):
        c0 = c0[2:]
    if c1 and c1.lower() != "nan":
        mtm = normalize_mtm(c0)
        serial = normalize_serial(c1)
        if mtm and serial:
            return mtm, serial
    blob = c0.upper().replace(" ", "")
    if len(blob) > 10 and LENOVO_MTM_RE.match(blob[:10]):
        serial = normalize_serial(blob[10:])
        if serial:
            return blob[:10], serial
    m = SCAN_SERIAL_TAIL_RE.search(blob)
    if m:
        serial = normalize_serial(m.group("serial"))
        mtm_part = blob[: m.start()]
        if mtm_part and serial:
            return normalize_mtm(mtm_part), serial
    if LENOVO_MTM_RE.match(blob):
        return blob, ""
    return "", ""


def looks_like_portal_scan_stock(path: Path) -> bool:
    """SCANNED STOCK FOR PORTAL export: 2 columns, often 1S{MTM}{SN} in one cell."""
    name = path.name.lower()
    if all(marker in name for marker in PORTAL_SCAN_NAME_MARKERS):
        return True
    try:
        df = read_tabular(path, nrows=40)
    except Exception:
        return False
    if df.shape[1] != 2:
        return False
    hits = 0
    for _, row in df.iterrows():
        c0 = str(row.iloc[0] if pd.notna(row.iloc[0]) else "").strip()
        if not c0 or _is_portal_scan_noise(c0):
            continue
        cu = c0.upper()
        if cu.startswith("1S") or re.match(r"^\d{2}[A-Z0-9]{8}(?:PC|PF|GM|R)", cu):
            hits += 1
    return hits >= 3


def load_portal_scan_stock(path: Path) -> pd.DataFrame:
    df = read_tabular(path)
    if df.shape[1] < 2:
        raise ValueError(f"Portal scan file needs 2 columns. Found: {list(df.columns)}")
    rows = []
    for _, row in df.iterrows():
        mtm, serial = parse_portal_scan_row(row.iloc[0], row.iloc[1] if len(row) > 1 else "")
        if mtm and serial:
            rows.append({"mtm": mtm, "serial": serial, "no_ssd": False, "mtm_raw": mtm})
    if not rows:
        raise ValueError("Portal scan file has no parseable laptop/desktop rows.")
    out = pd.DataFrame(rows)
    out = out.drop_duplicates(subset=["serial"], keep="first")
    return out.reset_index(drop=True)

def normalize_manufacturer(raw: str, mtm: str = "", model_hint: str = "") -> str:
    """Map Blancco / inferred text to LENOVO|DELL|HP|PANASONIC."""
    s = (raw or "").strip().upper()
    if "DELL" in s:
        return "DELL"
    if "LENOVO" in s:
        return "LENOVO"
    if "HP" in s or "HEWLETT" in s:
        return "HP"
    if "PANASONIC" in s:
        return "PANASONIC"
    return infer_manufacturer(mtm, model_hint)

def infer_manufacturer(mtm: str, model_hint: str = "") -> str:
    """Guess manufacturer from MTM / model text for Odoo import."""
    mtm_u = (mtm or "").strip().upper()
    hint = (model_hint or "").strip().upper()
    combined = f"{mtm_u} {hint}"
    if re.match(r"^\d{2}[A-Z0-9]{8}$", mtm_u) or mtm_u.startswith(("10", "20")):
        return "LENOVO"
    if (
        "DELL" in combined
        or "LATITUDE" in combined
        or "OPTIPLEX" in combined
        or hint.startswith("LATITUDE")
    ):
        return "DELL"
    if "PANASONIC" in combined or mtm_u.startswith(("CF-", "FZ-")):
        return "PANASONIC"
    if "ASUS" in combined:
        return "ASUS"
    if "HP" in combined or "#ABG" in combined:
        return "HP"
    return "LENOVO"

def parse_size_gb(text) -> str:
    if text is None or (isinstance(text, float) and pd.isna(text)):
        return ""
    s = str(text).strip()
    m = re.search(r"(\d+)", s)
    return m.group(1) if m else ""


def parse_battery_percents(text) -> list[int]:
    """Parse one or more battery health % values (e.g. '91;92' dual-battery)."""
    if text is None or (isinstance(text, float) and pd.isna(text)):
        return []
    s = str(text).strip()
    if not s:
        return []
    out = []
    for part in re.split(r"[;/,|]+", s):
        m = re.search(r"(\d+)", part.strip())
        if not m:
            continue
        n = int(m.group(1))
        if 0 < n <= 200:
            out.append(n)
    return out


def battery_display_label(percents: list[int]) -> str:
    if not percents:
        return "Unknown"
    return " / ".join(f"{p}%" for p in percents)


def battery_tier_label(percents: list[int]) -> str:
    """70%+ when min valid % >= 70; missing/unknown -> Under 70%."""
    if not percents:
        return "Under 70%"
    if min(percents) >= BATTERY_TIER_THRESHOLD:
        return "70%+"
    return "Under 70%"


def battery_tier_code(tier_label: str) -> str:
    return "BT70" if tier_label == "70%+" else "BTU70"


def is_laptop_product(model_name: str, mtm: str) -> bool:
    name = (model_name or "").lower()
    desktop_kw = (
        "thinkcentre", "thinkstation", "optiplex", "prodesk",
        "elitedesk", "tiny", " sff", "desktop", "workstation",
        "m70", "m73", "m910", "m920", "m93",
    )
    if any(k in name for k in desktop_kw):
        return False
    return True


def _pick_battery_col(col_map: dict[str, str]) -> str | None:
    disk_col = _pick_col(col_map, "disk capacity", "disk capacit")
    for name in (
        "battery 2 capacity",
        "battery 1 capacity",
        "battery capacity",
        "battery health",
        "battery (%)",
        "battery percentage",
        "battery",
        "capacity",
    ):
        col = _pick_col(col_map, name)
        if not col or col == disk_col:
            continue
        return col
    return None


def _battery_raw_from_row(row, battery_col: str | None, extra_battery_cols: list[str]) -> str:
    parts = []
    for col in extra_battery_cols:
        val = str(row.get(col, "") or "").strip()
        if val:
            parts.append(val)
    if battery_col:
        val = str(row.get(battery_col, "") or "").strip()
        if val:
            if parts:
                parts.insert(0, val)
            else:
                return val
    if not parts:
        return ""
    merged = []
    for part in parts:
        merged.extend(parse_battery_percents(part))
    if not merged:
        return ""
    return ";".join(str(p) for p in merged)

def parse_name_attrs(name: str) -> dict:
    n = (name or "").strip()
    low = n.lower()
    touch = ""
    if "non-touch" in low or "non touch" in low:
        touch = "No"
    elif "touch" in low:
        touch = "Yes"
    wan = ""
    if re.search(r"\bwan\b", low):
        wan = "Yes"
    elif touch == "Yes":
        wan = "No"
    gen = ""
    m = GEN_RE.search(n)
    if m:
        gen = m.group(1)
    clean = re.sub(r"^lenovo\s+", "", n, flags=re.I).strip()
    # WD receipt name — strip generation suffix; Gen comes from Blancco System version.
    clean = re.sub(r"\s*Gen(?:eration)?\s*\d+\w*\s*$", "", clean, flags=re.I).strip()
    return {"model_name": clean, "touch": touch, "wan": wan, "generation": gen}

def parse_generation_from_system_version(system_version: str) -> str:
    """Gen 1 / Gen 2i / Gen 3 from Blancco System version (authoritative)."""
    if not system_version:
        return ""
    m = SYSVER_GEN_RE.search(str(system_version))
    return m.group(1) if m else ""

def _looks_like_valid_family_label(text: str) -> bool:
    if not text or len(str(text).strip()) < 8:
        return False
    s = str(text).strip()
    if re.fullmatch(r"[A-Z0-9]{1,4}", s, re.I):
        return False
    low = s.lower()
    if "kbc" in low or "version" in low:
        return False
    return bool(
        re.search(
            r"thinkpad|thinkcentre|latitude|toughbook|optiplex|panasonic|elitebook|elitedesk|prodesk|cf-?\d|fz-?\d|dell",
            low,
            re.I,
        )
    )

def derive_product_name(
    system_version: str, model_name: str, mtm: str, generation: str = ""
) -> str:
    """Shop product name = Blancco System version when available (e.g. ThinkPad T14s Gen 2i)."""
    sv = str(system_version or "").strip()
    if sv and _looks_like_valid_family_label(sv):
        return sv
    gen = (generation or "").strip() or parse_generation_from_system_version(sv)
    if model_name and _looks_like_valid_family_label(model_name):
        base = re.sub(
            r"\s*Gen(?:eration)?\s*\d+\w*\s*$", "", str(model_name).strip(), flags=re.I
        ).strip()
        if gen:
            return f"{base} Gen {gen}"
        return base
    return mtm

_FILE_DIALOG_TYPES = [
    ("CSV files", "*.csv"),
    ("Excel files", "*.xlsx"),
    ("All supported", "*.csv;*.xlsx"),
]

def is_output_file(path: Path) -> bool:
    name_lower = path.name.lower()
    return any(name_lower.startswith(p) for p in OUTPUT_SKIP_PREFIXES)

def read_tabular(path: Path, nrows: int | None = None) -> pd.DataFrame:
    kwargs = {"dtype": str}
    if nrows is not None:
        kwargs["nrows"] = nrows
    if path.suffix.lower() == ".csv":
        return pd.read_csv(path, **kwargs)
    if path.suffix.lower() == ".xlsx":
        return pd.read_excel(path, sheet_name=0, **kwargs)
    raise ValueError(f"Unsupported file type '{path.suffix}'. Use .csv or .xlsx.")

def column_names(path: Path) -> list[str]:
    df = read_tabular(path, nrows=0)
    return [str(c).strip() for c in df.columns]

def _col_lookup(cols: list[str]) -> dict[str, str]:
    return {c.lower(): c for c in cols}

def _pick_col(col_map: dict[str, str], *names: str) -> str | None:
    for name in names:
        if name.lower() in col_map:
            return col_map[name.lower()]
    return None

def looks_like_product_list(cols: list[str]) -> bool:
    """Simple list: model/MTM + serial (typically 2–3 columns)."""
    if len(cols) > PRODUCT_LIST_MAX_COLS:
        return False
    col_map = _col_lookup(cols)
    has_mtm = _pick_col(col_map, "device model", "model mtm", "mtm", "model") is not None
    has_serial = _pick_col(col_map, "serial", "serial no.", "serial no") is not None
    if has_mtm and has_serial:
        return True
    return len(cols) == 2


def is_product_list_file(path: Path) -> bool:
    if looks_like_portal_scan_stock(path):
        return True
    return classify_file(path) == "product"

def looks_like_blancco(cols: list[str]) -> bool:
    """Detailed Blancco export: serial + many spec columns."""
    col_map = _col_lookup(cols)
    has_serial = _pick_col(col_map, "system serial", "serial", "system serial number") is not None
    if not has_serial:
        return False
    markers = (
        "creation date",
        "system version",
        "system model",
        "cpu model",
        "disk capacity",
        "physical_memory",
        "physical memory",
        "motherboard test status",
        "video card model",
    )
    marker_hits = sum(1 for m in markers if m in col_map)
    return len(cols) >= BLANCCO_MIN_COLS or marker_hits >= 2

def classify_file(path: Path) -> str:
    """Return 'product', 'blancco', or 'unknown'."""
    cols = column_names(path)
    is_product = looks_like_product_list(cols)
    is_blancco = looks_like_blancco(cols)
    if is_product and not is_blancco:
        return "product"
    if is_blancco and not is_product:
        return "blancco"
    if is_product and is_blancco:
        return "product" if len(cols) <= PRODUCT_LIST_MAX_COLS else "blancco"
    return "unknown"

def file_type_label(kind: str) -> str:
    return {"product": "Product list", "blancco": "Blancco report"}.get(kind, "Unknown file")

def scan_folder(folder: Path) -> tuple[Path | None, Path | None]:
    """Pick newest product list + newest Blancco report by column layout."""
    products: list[Path] = []
    blanccos: list[Path] = []
    for path in folder.iterdir():
        if not path.is_file() or path.suffix.lower() not in SUPPORTED_EXTENSIONS:
            continue
        if is_output_file(path):
            continue
        try:
            if is_product_list_file(path):
                products.append(path)
            elif classify_file(path) == "blancco":
                blanccos.append(path)
        except Exception:
            continue
    product = max(products, key=lambda p: p.stat().st_mtime) if products else None
    blancco = max(blanccos, key=lambda p: p.stat().st_mtime) if blanccos else None
    return product, blancco

def master_label(path: Path) -> str:
    if looks_like_portal_scan_stock(path):
        return "Portal scan stock"
    return file_type_label(classify_file(path))

def _ask_open_file(title: str, folder: Path) -> Path | None:
    import tkinter as tk
    from tkinter import filedialog

    root = tk.Tk()
    root.withdraw()
    root.attributes("-topmost", True)
    selected = filedialog.askopenfilename(
        title=title,
        initialdir=str(folder),
        filetypes=_FILE_DIALOG_TYPES,
    )
    root.destroy()
    if not selected:
        return None
    return Path(selected)

def pick_files_interactive(folder: Path) -> tuple[Path, Path]:
    product, blancco = scan_folder(folder)
    force_pick = "--pick" in sys.argv

    if "--yes" in sys.argv or "--auto" in sys.argv:
        if not product or not blancco:
            raise FileNotFoundError(
                "Auto mode: need one product list (model + serial) and one Blancco report in "
                + str(folder)
            )
        print("Auto merge:", product.name, "+", blancco.name)
        return product, blancco

    import tkinter as tk
    from tkinter import messagebox

    root = tk.Tk()
    root.withdraw()
    root.attributes("-topmost", True)

    if product and blancco and not force_pick:
        msg = (
            "Auto-detected by file layout:\n\n"
            f"Product list:\n  {product.name}\n\n"
            f"Blancco report:\n  {blancco.name}\n\n"
            "Yes = merge with these files\n"
            "No = choose different files\n"
            "Cancel = abort"
        )
        choice = messagebox.askyesnocancel("Re-Ware merge", msg)
        root.destroy()
        if choice is True:
            return product, blancco
        if choice is None:
            sys.exit(0)
        product = None
        blancco = None
    else:
        root.destroy()
        hints = ["Could not auto-detect both files by column layout."]
        if product:
            hints.append(f"Product list found: {product.name}")
        if blancco:
            hints.append(f"Blancco report found: {blancco.name}")
        hints.append(
            "\nYou will pick two files:\n"
            "  1) Product list — usually 2 columns (model/MTM + serial)\n"
            "  2) Blancco report — many columns (CPU, disk, etc.)"
        )
        messagebox.showinfo("Re-Ware merge", "\n".join(hints))

    if not product or force_pick:
        product = _ask_open_file(
            "Step 1/2 — Product list (.csv or .xlsx: model/MTM + serial)",
            folder,
        )
        if not product:
            sys.exit(0)

    if not blancco or force_pick:
        blancco = _ask_open_file(
            "Step 2/2 — Blancco report (.csv or .xlsx: detailed export)",
            folder,
        )
        if not blancco:
            sys.exit(0)

    return product, blancco

def validate_merge_inputs(product_path: Path, blancco_path: Path) -> None:
    if not product_path.is_file():
        raise FileNotFoundError(f"Product list not found:\n{product_path}")
    if not blancco_path.is_file():
        raise FileNotFoundError(f"Blancco report not found:\n{blancco_path}")
    if product_path.resolve() == blancco_path.resolve():
        raise ValueError("Product list and Blancco report must be two different files.")
    if product_path.suffix.lower() not in SUPPORTED_EXTENSIONS:
        raise ValueError(f"Product list must be .csv or .xlsx, not '{product_path.suffix}'.")
    if blancco_path.suffix.lower() not in SUPPORTED_EXTENSIONS:
        raise ValueError(f"Blancco report must be .csv or .xlsx, not '{blancco_path.suffix}'.")
    if is_output_file(product_path):
        raise ValueError(
            f"'{product_path.name}' looks like a previous merge output.\n"
            "Pick the simple product list (model + serial), not MERGED import-ready."
        )
    if is_output_file(blancco_path):
        raise ValueError(
            f"'{blancco_path.name}' looks like a previous merge output.\n"
            "Pick the Blancco export file instead."
        )

    product_kind = classify_file(product_path)
    blancco_kind = classify_file(blancco_path)

    if product_kind == "blancco" and (blancco_kind == "product" or is_product_list_file(blancco_path)):
        raise ValueError(
            "Files look swapped.\n\n"
            f"First file '{product_path.name}' looks like a Blancco report.\n"
            f"Second file '{blancco_path.name}' looks like a product list.\n\n"
            "Run again and pick product list first, then Blancco report."
        )
    if not is_product_list_file(product_path):
        cols = column_names(product_path)
        raise ValueError(
            f"'{product_path.name}' does not look like a product list.\n"
            f"Found {len(cols)} column(s): {', '.join(cols[:6])}{'...' if len(cols) > 6 else ''}\n\n"
            "Expected model/MTM + serial, or SCANNED STOCK FOR PORTAL (1S scan rows)."
        )
    if blancco_kind != "blancco":
        cols = column_names(blancco_path)
        raise ValueError(
            f"'{blancco_path.name}' does not look like a Blancco report.\n"
            f"Found {len(cols)} column(s): {', '.join(cols[:6])}{'...' if len(cols) > 6 else ''}\n\n"
            "Expected the detailed Blancco export (serial + CPU, disk, RAM, etc.)."
        )

def load_product_list(path: Path) -> pd.DataFrame:
    if looks_like_portal_scan_stock(path):
        return load_portal_scan_stock(path)
    df = read_tabular(path)
    df.columns = [str(c).strip() for c in df.columns]
    col_map = _col_lookup(list(df.columns))
    mtm_col = _pick_col(col_map, "device model", "model mtm", "mtm", "model")
    serial_col = _pick_col(col_map, "serial", "serial no.", "serial no")
    has_ssd_col = _pick_col(col_map, "has_ssd", "has ssd")

    if (not mtm_col or not serial_col) and len(df.columns) == 2:
        mtm_col, serial_col = df.columns[0], df.columns[1]
    if not mtm_col or not serial_col:
        raise ValueError(
            f"Product list needs model/MTM and serial columns. Found: {list(df.columns)}"
        )

    cols = [mtm_col, serial_col]
    if has_ssd_col:
        cols.append(has_ssd_col)
    out = df[cols].copy()
    out.columns = ["mtm_raw", "serial_raw"] + (["has_ssd_raw"] if has_ssd_col else [])
    out["mtm"] = out["mtm_raw"].map(normalize_mtm)
    out["serial"] = out["serial_raw"].map(normalize_serial)
    if has_ssd_col:
        out["no_ssd"] = out["has_ssd_raw"].fillna("").str.strip().str.upper().eq("NO SSD")
    else:
        out["no_ssd"] = False
    out = out[(out["serial"] != "") & (out["mtm"] != "")]
    if out.empty:
        raise ValueError("Product list has no valid rows with both model/MTM and serial.")
    out = out.drop_duplicates(subset=["serial"], keep="first")
    return out[["mtm", "serial", "no_ssd", "mtm_raw"]].reset_index(drop=True)

def load_master(path: Path) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    """Load product list; optional sheet-2 metadata from legacy delivery receipts."""
    if not is_product_list_file(path):
        raise ValueError(
            f"'{path.name}' is not a product list. "
            "Use the simple file with model/MTM + serial columns, or SCANNED STOCK FOR PORTAL."
        )
    wd = load_product_list(path)
    uncollected_df = pd.DataFrame(columns=["serial", "uncollected"])
    sheet2_names = pd.DataFrame()
    if path.suffix.lower() == ".xlsx":
        uncollected_df, sheet2_names = load_receipt_sheet2_meta(path)
    return wd, uncollected_df, sheet2_names

def load_receipt_sheet2_meta(path: Path) -> tuple[pd.DataFrame, pd.DataFrame]:
    """Returns (serial->comments/uncollected, rows for MTM LUT generation)."""
    try:
        raw = pd.read_excel(path, sheet_name=1, header=None, dtype=str)
    except Exception:
        return pd.DataFrame(columns=["serial", "uncollected"]), pd.DataFrame()

    header_row = None
    for i, row in raw.iterrows():
        cells = [str(x).strip().lower() for x in row.tolist() if pd.notna(x)]
        if "serial no." in cells or "serial no" in cells:
            header_row = i
            break
    if header_row is None:
        return pd.DataFrame(columns=["serial", "uncollected"]), pd.DataFrame()

    df = pd.read_excel(path, sheet_name=1, header=int(header_row), dtype=str)
    df.columns = [str(c).strip() for c in df.columns]
    col_map = {c.lower(): c for c in df.columns}
    serial_col = col_map.get("serial no.") or col_map.get("serial no")
    name_col = col_map.get("name")
    comments_col = col_map.get("comments")
    if not serial_col:
        return pd.DataFrame(columns=["serial", "uncollected"]), pd.DataFrame()

    meta = df[[serial_col] + ([comments_col] if comments_col else []) + ([name_col] if name_col else [])].copy()
    meta["serial"] = meta[serial_col].map(normalize_serial)
    meta = meta[meta["serial"] != ""]
    if comments_col:
        meta["uncollected"] = meta[comments_col].fillna("").str.lower().str.contains("uncollected")
    else:
        meta["uncollected"] = False
    if name_col:
        meta["name"] = meta[name_col].fillna("")

    uncollected_df = meta[["serial", "uncollected"]].drop_duplicates("serial")
    lut_rows = meta[["serial", "name"]].drop_duplicates("serial") if name_col else pd.DataFrame()
    return uncollected_df, lut_rows

def normalize_sku(value) -> str:
    """System SKU number (HP / Panasonic) — preserve CF-54… / #ABG style codes."""
    return normalize_mtm(value)

def is_sku_title_vendor(manufacturer: str, mtm: str = "", model_hint: str = "") -> bool:
    """HP / Panasonic / Dell: Blancco *system model* = title, *system SKU* = product code.

  Lenovo is the exception (system *version* = title, system *model* = MTM).
    """
    if is_lenovo_style_mtm(mtm) or is_lenovo_style_mtm(model_hint):
        return False
    mfr = (manufacturer or "").strip().upper()
    if mfr in ("HP", "PANASONIC", "DELL"):
        return True
    blob = f"{mtm or ''} {model_hint or ''}".upper()
    return bool(
        re.search(
            r"CF-?\d|FZ-?\d|TOUGHBOOK|#ABG|T1D\d|LATITUDE|OPTIPLEX|ELITEDESK|PRODESK",
            blob,
        )
        or blob.startswith(("CF", "FZ", "DELL"))
    )


def is_hp_or_panasonic(manufacturer: str, mtm: str = "", model_hint: str = "") -> bool:
    """Backward-compatible alias — prefer :func:`is_sku_title_vendor`."""
    return is_sku_title_vendor(manufacturer, mtm, model_hint)

def is_lenovo_style_mtm(mtm: str) -> bool:
    mtm_u = (mtm or "").strip().upper()
    return bool(re.match(r"^\d{2}[A-Z0-9]{8}$", mtm_u) or mtm_u.startswith(("10", "20", "30")))

def build_blancco_indexes(blancco: pd.DataFrame):
    """Index Blancco rows by device serial, system SKU number, and system model."""
    by_serial = blancco.set_index("serial", drop=False)
    sku_rows = blancco[blancco["sku_number"].astype(str).str.strip() != ""]
    by_sku = (
        sku_rows.drop_duplicates(subset=["sku_number"], keep="first").set_index(
            "sku_number", drop=False
        )
        if not sku_rows.empty
        else pd.DataFrame().set_index(pd.Index([], name="sku_number"))
    )
    model_rows = blancco[blancco["blancco_mtm"].astype(str).str.strip() != ""]
    by_model = (
        model_rows.drop_duplicates(subset=["blancco_mtm"], keep="first").set_index(
            "blancco_mtm", drop=False
        )
        if not model_rows.empty
        else pd.DataFrame().set_index(pd.Index([], name="blancco_mtm"))
    )
    return by_serial, by_sku, by_model

def resolve_blancco_row(by_serial, by_sku, by_model, receipt_serial: str, receipt_mtm: str):
    """Match receipt row to Blancco.

    Lenovo: receipt serial ≈ Blancco *system serial*.
    HP / Panasonic / Dell: receipt may list *system SKU* in the serial or MTM column —
    join via *system SKU number*, then use Blancco *system serial* as the device SN.
    """
    for key, idx, kind in (
        (receipt_serial, by_serial, "system_serial"),
        (receipt_serial, by_sku, "system_sku"),
        (receipt_mtm, by_sku, "mtm_as_sku"),
    ):
        if not key or idx.empty or key not in idx.index:
            continue
        row = idx.loc[key]
        if isinstance(row, pd.DataFrame):
            row = row.iloc[0]
        return row, kind
    # MTM/model fallback: only for SKU-style receipts without a unit serial (HP/Dell/Panasonic).
    # Never match Lenovo by MTM alone — would assign the wrong device SN/specs.
    if receipt_mtm and receipt_mtm in by_model.index:
        if receipt_serial and is_lenovo_style_mtm(receipt_mtm):
            return None, ""
        row = by_model.loc[receipt_mtm]
        if isinstance(row, pd.DataFrame):
            row = row.iloc[0]
        return row, "system_model"
    return None, ""

def output_device_serial(receipt_serial: str, bl_row, match_kind: str) -> str:
    """Odoo / CSV Serial column = real unit SN from Blancco when known."""
    if bl_row is None:
        return receipt_serial
    device = str(bl_row.get("serial", "") or "").strip().upper()
    if match_kind in ("system_sku", "mtm_as_sku") and device:
        return device
    if match_kind == "system_serial" and device:
        return device
    if match_kind == "system_model":
        return receipt_serial
    return receipt_serial

def output_mtm(receipt_mtm: str, bl_row, manufacturer: str) -> str:
    """MTM/SKU for Odoo.

    Lenovo: Blancco *system model* (20WN… MTM).
    HP / Panasonic / Dell: Blancco *system SKU number* — receipt Device Model column;
    *system model* is the shop title only.
    """
    if bl_row is None:
        return receipt_mtm
    sku = normalize_sku(str(bl_row.get("sku_number", "") or ""))
    bl_mtm = normalize_mtm(str(bl_row.get("blancco_mtm", "") or ""))
    if is_sku_title_vendor(manufacturer, receipt_mtm, sku or bl_mtm):
        if sku and not sku.upper().startswith("LENOVO_MT_"):
            return sku
        return receipt_mtm
    if bl_mtm and not is_lenovo_style_mtm(receipt_mtm):
        return bl_mtm
    return receipt_mtm

def load_blancco(path: Path) -> pd.DataFrame:
    if classify_file(path) != "blancco":
        cols = column_names(path)
        raise ValueError(
            f"'{path.name}' does not look like a Blancco report ({len(cols)} columns). "
            "Expected serial + detailed specs (CPU, disk, RAM, etc.)."
        )
    df = read_tabular(path)
    df.columns = [str(c).strip() for c in df.columns]
    col_map = _col_lookup(list(df.columns))

    serial_col = _pick_col(col_map, "system serial", "serial", "system serial number")
    sku_col = _pick_col(
        col_map,
        "system sku number",
        "system sku",
        "system sku no.",
        "system sku no",
        "sku number",
    )
    if not serial_col:
        raise ValueError(f"Blancco file needs a serial column. Found: {list(df.columns)}")

    battery_col = _pick_battery_col(col_map)
    battery_extra_cols = [
        c
        for c in (
            _pick_col(col_map, "battery 1 capacity"),
            _pick_col(col_map, "battery 2 capacity"),
        )
        if c
    ]

    mapping = {
        "blancco_date": _pick_col(col_map, "creation date"),
        "blancco_title": _pick_col(col_map, "system version"),
        "blancco_mtm": _pick_col(col_map, "system model"),
        "cpu": _pick_col(col_map, "cpu model", "cpu"),
        "disk_capacity": _pick_col(col_map, "disk capacity", "disk capacit"),
        "ssd_type": _pick_col(col_map, "disk interface type"),
        "ram": _pick_col(col_map, "physical_memory", "physical memory", "ram"),
        "gpu": _pick_col(col_map, "video card model", "gpu"),
        "mobo_status": _pick_col(col_map, "motherboard test status", "cmos condition"),
        "manufacturer": _pick_col(col_map, "system manufacturer", "manufacturer"),
    }

    out = pd.DataFrame()
    out["serial"] = df[serial_col].map(normalize_serial)
    out["sku_number"] = df[sku_col].map(normalize_sku) if sku_col else ""
    for key, col in mapping.items():
        out[key] = df[col] if col else ""
    out["battery"] = df.apply(
        lambda row: _battery_raw_from_row(row, battery_col, battery_extra_cols),
        axis=1,
    )
    if "blancco_mtm" in out.columns:
        out["blancco_mtm"] = out["blancco_mtm"].map(normalize_mtm)
    out = out[out["serial"] != ""]
    out["_duplicate_count"] = out.groupby("serial")["serial"].transform("count")

    if "blancco_date" in out.columns and out["blancco_date"].notna().any():
        out = out.sort_values("blancco_date", ascending=False)
    out = out.drop_duplicates(subset=["serial"], keep="first")
    return out.reset_index(drop=True)

def ensure_mtm_lookup(receipt_path: Path, sheet1: pd.DataFrame, sheet2_names: pd.DataFrame) -> pd.DataFrame:
    if MTM_LUT_FILE.is_file():
        lut = pd.read_csv(MTM_LUT_FILE, dtype=str)
        lut["mtm"] = lut["mtm"].map(normalize_mtm)
        return lut

    if sheet2_names.empty:
        lut = pd.DataFrame(columns=["mtm", "model_name", "touch", "wan", "generation"])
        lut.to_csv(MTM_LUT_FILE, index=False)
        return lut

    joined = sheet2_names.merge(sheet1[["serial", "mtm"]], on="serial", how="inner")
    rows = []
    for mtm, group in joined.groupby("mtm"):
        names = group["name"].dropna().astype(str)
        name = names.mode().iloc[0] if len(names) else ""
        attrs = parse_name_attrs(name)
        rows.append({"mtm": mtm, **attrs})
    lut = pd.DataFrame(rows)
    lut.to_csv(MTM_LUT_FILE, index=False)
    return lut

def classify_row(wd_row, blancco_row, uncollected: bool, no_ssd: bool = False) -> tuple[str, str]:
    if uncollected:
        return "FAILED", "Uncollected (not received)"
    if blancco_row is None:
        return "FAILED", "Serial not found in Blancco"
    if int(blancco_row.get("_duplicate_count", 1) or 1) > 1:
        pass  # already deduped; note in reason if needed
    cpu = str(blancco_row.get("cpu", "") or "").strip()
    disk = str(blancco_row.get("disk_capacity", "") or "").strip()
    if no_ssd and cpu:
        return "SUCCESS", ""
    if not cpu and not disk:
        return "FAILED", "Serial in Blancco but specs empty"
    return "SUCCESS", ""

def merge_data(
    wd: pd.DataFrame,
    blancco: pd.DataFrame,
    lut: pd.DataFrame,
    uncollected_map: dict[str, bool],
) -> pd.DataFrame:
    by_serial, by_sku, by_model = build_blancco_indexes(blancco)
    lut_idx = lut.set_index("mtm", drop=False) if not lut.empty else {}

    rows = []
    for _, wd_row in wd.iterrows():
        receipt_serial = wd_row["serial"]
        receipt_mtm = wd_row["mtm"]
        bl, match_kind = resolve_blancco_row(
            by_serial, by_sku, by_model, receipt_serial, receipt_mtm
        )

        uncollected = uncollected_map.get(receipt_serial, False)
        no_ssd = bool(wd_row.get("no_ssd", False))
        status, reason = classify_row(wd_row, bl, uncollected, no_ssd=no_ssd)

        lut_row = lut_idx.loc[receipt_mtm] if receipt_mtm in getattr(lut_idx, "index", []) else None
        if lut_row is not None and isinstance(lut_row, pd.DataFrame):
            lut_row = lut_row.iloc[0]

        model_name = ""
        touch = wan = ""
        if lut_row is not None:
            model_name = str(lut_row.get("model_name", "") or "")
            model_name = re.sub(
                r"\s*Gen(?:eration)?\s*\d+\w*\s*$", "", model_name, flags=re.I
            ).strip()
            touch = str(lut_row.get("touch", "") or "")
            wan = str(lut_row.get("wan", "") or "")

        system_version = str(bl.get("blancco_title", "") or "") if bl is not None else ""
        bl_mtm = str(bl.get("blancco_mtm", "") or "") if bl is not None else ""
        mtm_raw = str(wd_row.get("mtm_raw", "") or "").strip()
        bl_mfr = str(bl.get("manufacturer", "") if bl is not None else "")
        manufacturer = normalize_manufacturer(
            bl_mfr,
            receipt_mtm,
            model_name or mtm_raw or bl_mtm,
        )
        mtm = output_mtm(receipt_mtm, bl, manufacturer)
        device_serial = output_device_serial(receipt_serial, bl, match_kind)

        if is_sku_title_vendor(manufacturer, mtm, model_name or bl_mtm):
            if bl_mtm:
                model_name = bl_mtm
            elif mtm_raw:
                model_name = mtm_raw
            system_version = ""
        elif not model_name and system_version and _looks_like_valid_family_label(system_version):
            model_name = re.sub(
                r"\s*Gen(?:eration)?\s*\d+\w*\s*$", "", system_version, flags=re.I
            ).strip()
        elif not model_name and mtm_raw and not is_lenovo_style_mtm(receipt_mtm):
            model_name = mtm_raw

        gen = parse_generation_from_system_version(
            str(bl.get("blancco_title", "") or "") if bl is not None else ""
        )
        if not gen and lut_row is not None:
            gen = str(lut_row.get("generation", "") or "").strip()
        if is_sku_title_vendor(manufacturer, mtm, model_name):
            series = model_name or derive_product_name("", model_name, mtm, gen)
        else:
            series = derive_product_name(
                str(bl.get("blancco_title", "") or "") if bl is not None else "",
                model_name,
                mtm,
                gen,
            )

        ssd_type = str(bl.get("ssd_type", "") if bl is not None else "")
        ssd_size = parse_size_gb(bl.get("disk_capacity", "") if bl is not None else "")
        if no_ssd:
            ssd_type = ""
            ssd_size = ""

        battery_raw = str(bl.get("battery", "") if bl is not None else "")
        battery_percents = parse_battery_percents(battery_raw)
        battery_display = battery_display_label(battery_percents)
        battery_tier = battery_tier_label(battery_percents)
        tier_code = battery_tier_code(battery_tier)
        if is_laptop_product(model_name, mtm):
            shop_sku = f"{mtm}-{tier_code}"
        else:
            shop_sku = mtm

        rows.append(
            {
                "Serial": device_serial,
                "MTM": mtm,
                "Model name": model_name,
                "System version": system_version,
                "Series": series,
                "Touch": touch,
                "WAN": wan,
                "Generation": gen,
                "CPU": str(bl.get("cpu", "") if bl is not None else ""),
                "RAM (GB)": parse_size_gb(bl.get("ram", "") if bl is not None else ""),
                "SSD type": ssd_type,
                "SSD size (GB)": ssd_size,
                "Battery (%)": battery_raw,
                "Battery display": battery_display,
                "Battery tier": battery_tier,
                "Shop SKU": shop_sku,
                "GPU": str(bl.get("gpu", "") if bl is not None else ""),
                "Mobo status": str(bl.get("mobo_status", "") if bl is not None else ""),
                "Blancco date": str(bl.get("blancco_date", "") if bl is not None else ""),
                "Manufacturer": manufacturer,
                "Price": "",
                "Status": status,
                "Failure reason": reason,
            }
        )
    return pd.DataFrame(rows, columns=OUTPUT_COLUMNS)

def build_analysis(merged: pd.DataFrame) -> pd.DataFrame:
    total = len(merged)
    success = int((merged["Status"] == "SUCCESS").sum())
    failed = total - success
    rate = f"{(success / total * 100):.1f}%" if total else "0%"
    unique_serials = int(merged["Serial"].nunique())
    dup_serials = int(total - unique_serials)

    lines = [
        ["Metric", "Value"],
        ["Master rows", total],
        ["Unique serials in output", unique_serials],
        ["Duplicate serial rows", dup_serials],
        ["Matched (SUCCESS)", success],
        ["Failed", failed],
        ["Match rate", rate],
        ["", ""],
        ["Failure breakdown", "Count"],
    ]
    if failed:
        for reason, cnt in merged.loc[merged["Status"] == "FAILED", "Failure reason"].value_counts().items():
            lines.append([reason, int(cnt)])
    else:
        lines.append(["(none)", 0])
    lines.append(["", ""])
    lines.append(["Failed serials (for investigation)", "MTM"])
    for _, r in merged.loc[merged["Status"] == "FAILED"].iterrows():
        lines.append([r["Serial"], r["MTM"]])

    return pd.DataFrame(lines)

def write_output(merged: pd.DataFrame, analysis: pd.DataFrame, out_path: Path) -> None:
    failed = merged.loc[merged["Status"] == "FAILED"]
    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        merged.to_excel(writer, sheet_name="Devices", index=False)
        if not failed.empty:
            failed.to_excel(writer, sheet_name="Failed", index=False)
        analysis.to_excel(writer, sheet_name="Analysis", index=False, header=False)

    wb = load_workbook(out_path)
    ws = wb["Devices"]
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = Font(bold=True)

    status_col = OUTPUT_COLUMNS.index("Status") + 1
    for row_idx in range(2, ws.max_row + 1):
        if ws.cell(row=row_idx, column=status_col).value == "FAILED":
            for col_idx in range(1, len(OUTPUT_COLUMNS) + 1):
                ws.cell(row=row_idx, column=col_idx).fill = FAIL_FILL

    if "Failed" in wb.sheetnames:
        ws_fail = wb["Failed"]
        for cell in ws_fail[1]:
            cell.fill = HEADER_FILL
            cell.font = Font(bold=True)
        for row_idx in range(2, ws_fail.max_row + 1):
            for col_idx in range(1, len(OUTPUT_COLUMNS) + 1):
                ws_fail.cell(row=row_idx, column=col_idx).fill = FAIL_FILL

    wb.save(out_path)

def failure_summary_text(merged: pd.DataFrame) -> str:
    failed = merged[merged["Status"] == "FAILED"]
    if failed.empty:
        return ""
    lines = ["\nFailure breakdown:"]
    for reason, cnt in failed["Failure reason"].value_counts().items():
        lines.append(f"  {reason}: {int(cnt)}")
    lines.append("  (full failed list is in the Analysis sheet of the .xlsx)")
    return "\n".join(lines)

def show_result_popup(summary: str, out_path: Path) -> None:
    import tkinter as tk
    from tkinter import messagebox

    root = tk.Tk()
    root.withdraw()
    root.attributes("-topmost", True)
    messagebox.showinfo("Re-Ware merge complete", summary + f"\n\nSaved:\n{out_path}")
    root.destroy()

def main() -> int:
    folder = SCRIPT_DIR
    product_path, blancco_path = pick_files_interactive(folder)
    validate_merge_inputs(product_path, blancco_path)

    try:
        wd, uncollected_df, sheet2_names = load_master(product_path)
    except Exception as exc:
        raise ValueError(
            f"Could not read product list '{product_path.name}':\n{exc}"
        ) from exc
    try:
        blancco = load_blancco(blancco_path)
    except Exception as exc:
        raise ValueError(
            f"Could not read Blancco report '{blancco_path.name}':\n{exc}"
        ) from exc

    uncollected_map = dict(zip(uncollected_df["serial"], uncollected_df["uncollected"])) if not uncollected_df.empty else {}
    lut = ensure_mtm_lookup(product_path, wd, sheet2_names)

    merged = merge_data(wd, blancco, lut, uncollected_map)

    stamp = datetime.now().strftime("%Y-%m-%d")
    xlsx_path = folder / f"{OUTPUT_PREFIX} {stamp}.xlsx"
    csv_path = folder / f"{OUTPUT_PREFIX} {stamp}.csv"
    analysis = build_analysis(merged)
    write_output(merged, analysis, xlsx_path)
    merged.to_csv(csv_path, index=False, encoding="utf-8-sig")

    total = len(merged)
    success = int((merged["Status"] == "SUCCESS").sum())
    failed = total - success
    summary = (
        f"Product list ({master_label(product_path)}): {total} rows\n"
        f"SUCCESS (Blancco data pulled): {success}\n"
        f"FAILED: {failed}\n"
        f"Match rate: {(success/total*100):.1f}%" if total else "No rows"
    )
    summary += failure_summary_text(merged)
    summary += (
        f"\n\nSaved:\n"
        f"  {xlsx_path.name}  — review (red = failed, see Analysis sheet)\n"
        f"  {csv_path.name}  — all rows with Status / Failure reason\n"
        f"\nUpload SUCCESS rows to Odoo:\n"
        f"  Inventory -> Upload inventory CSV (after DB backup)"
    )
    if "--yes" in sys.argv or "--auto" in sys.argv:
        print(summary)
        print("Saved:", xlsx_path)
        print("Saved:", csv_path)
    else:
        show_result_popup(summary, xlsx_path)
    return 0 if failed == 0 else 2

if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except Exception as exc:
        import tkinter as tk
        from tkinter import messagebox

        root = tk.Tk()
        root.withdraw()
        root.attributes("-topmost", True)
        messagebox.showerror("Re-Ware merge error", str(exc))
        root.destroy()
        raise